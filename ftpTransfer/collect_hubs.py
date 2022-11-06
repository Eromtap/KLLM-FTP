# -*- coding: utf-8 -*-
"""
Created on Fri Sep  9 22:10:10 2022

@author: Mark Patmore



This version does not support empty strings for temp lists. Thus columns
    are not properly aligned with headings.


Module to open truck mileage spreadsheet and both Mcleod and TK reefer hr spreadsheets
    create list of data rows and convert to ODS
    

It's possible we may need to get data from Omnitracs because some of the new
    trucks dont seem to be reporting to Mcleod. Or, better yet, ask JJ to 
    fix this... Importing from Omni would add plenty of risk for errors/crashes ######################
    
    
    
    
THIS MODULE WILL MOSTLY BE USED FOR TESTING ##########################################################################    
    
    
    
    
    


"""

from pyexcel_ods import save_data
from collections import OrderedDict


from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import date

from definitions import ROOT_DIR

print(ROOT_DIR)


today = date.today()
today = today.strftime("%m/%d/%Y")

column_headings = [
                   "EQCompanyCode",
                   "EQFleetCode",
                   "EQUnitCode",
                   "EQType",
                   "EQVIN",
                   "MeterReadingDate",
                   "PrimaryMeterType",
                   "PrimaryMeterReading",
                   "PrimaryMeterLifeToDateUsage",
                   "SecondaryMeterType",
                   "SecondaryMeterReading",
                   "SecondaryMeterLifeToDateUsage",
                   "Latitude",
                   "Longitude"
                   ]

xlsx_rows = [column_headings]


## Collect data from xlsx and input into global xlsx_rows list
## Need to find out if strings are ok or if unit id's need to be int. ########################
# Probably strings since reefers will start with a 'U'
# Also need to find out if strings are OK for hub and dates ##################################

def collect_xlsx_data(filename, starting_row, unit_id_col, hub_col, date_col): 
    
    wb = load_workbook(filename)
    ws = wb.active    
    max_row = ws.max_row + 1
    
    
    ## Temp list to store row data
    temp_list = []
    
    for r in range(starting_row,max_row):
        unit_id = ws.cell(row=r, column=unit_id_col).value
        hub_reading = ws.cell(row=r, column=hub_col).value
        date = ws.cell(row=r, column=date_col).value
        
        if not hub_reading:
            hub_reading = ' '
            
        if not date:
            date = ' '
        
        
        temp_list.append(unit_id)
        temp_list.append(hub_reading)
        if date:
            temp_list.append(date[:-5])
        else:
            temp_list.append(date)
        
        
        ## Carve out for mc_trailers.xlsx because TK reefers don't 
        ##     track to Mcleod, so adding them would be stupid
        ## Instead, we will be adding them from report generated
        ##    by TK Tracking. Super annoying...
        ## This block will hopefully be removed when we get rid of all the
        ##    TK units in the not too distant future
        
        if  'mc_trailers.xlsx' in filename:
            if temp_list[0].startswith('5') and int(temp_list[0]) < 532111:
                temp_list = []
                continue
            else:
                if temp_list[0].startswith('5'):
                    temp_list[0] = f'U{temp_list[0]}'
                    xlsx_rows.append(temp_list)
        else:
            xlsx_rows.append(temp_list)
        
        temp_list = []
    

## Stupid func to get data from the TK Tracking sheet.
## Hope to be able to get rid of this eventually.  
## Carve out for '16 trailers that for some dumb reson still show up in
##     our equipment list.      

def collect_stupid_TK_data(filename, starting_row, unit_id_col, hub_col, date):        
    wb = load_workbook(filename)
    ws = wb.active    
    max_row = ws.max_row + 1
    
    
    ## Temp list to store row data
    temp_list = []
    
    for r in range(starting_row,max_row):
        unit_id = f'U{ws.cell(row=r, column=unit_id_col).value}'
        hub_reading = ws.cell(row=r, column=hub_col).value

        if unit_id.startswith("U5316"):
            continue
        else:
            temp_list.append(unit_id)
            temp_list.append(hub_reading)
            temp_list.append(date)
            
            xlsx_rows.append(temp_list)
            temp_list = []
        
        
        
        
        
def create_combined_ODS(filename):
    data = OrderedDict()

    data.update({"Sheet 1": xlsx_rows})
    
    
    save_data(filename, data)
        







        
        
        
collect_xlsx_data(f'{ROOT_DIR}\\sheets\\mc_trucks.xlsx',3,2,3,4)          
        
collect_stupid_TK_data(f'{ROOT_DIR}\\sheets\\tk_hrs.xlsx', 5, 1, 11, today)        
   
collect_xlsx_data(f'{ROOT_DIR}\\sheets\\mc_trailers.xlsx', 3, 2, 3, 4)   



create_combined_ODS(f'{ROOT_DIR}\\temp\\Cetarisimport.ods')




























