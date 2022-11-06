# -*- coding: utf-8 -*-
#! /home/markpatmore/QuestPrograms/QuestEnv/bin/python3


"""
Created on Fri Sep  9 22:10:10 2022

@author: Mark Patmore FMDJP




This is the version that includes temp list with empty strings to 
    align columns with headers




Module to open truck mileage spreadsheet and both Mcleod and TK reefer hr spreadsheets
    create list of data rows and convert to ODS
    

It's possible we may need to get data from Omnitracs because some of the new
    trucks dont seem to be reporting to Mcleod. Or, better yet, ask JJ to 
    fix this... Importing from Omni would add plenty of risk for errors/crashes ######################


"""

from pyexcel_ods import save_data
from collections import OrderedDict
from openpyxl import load_workbook


column_headings = [
                    "EqCompanyCode",
                    "EqFleetCode",
                    "EqUnitCode",
                    "EqType",
                    "EqVIN",
                    "MeterReadingDate",
                    "PrimaryMeterType",
                    "PrimaryMeterReading",
                    "PrimaryMeterLifeToDateUsage",
                    "SecondaryMeterType",
                    "SecondaryMeterReading",
                    "SecondaryMeterLifeToDateUsage",
                    "Latitude",
                    "Longitude"
                    ]

xlsx_rows = [column_headings]
# xlsx_rows = []


## Collect data from xlsx and input into global xlsx_rows list


def collect_xlsx_data(filename, starting_row, unit_id_col, unit_type, vin_col, hub_col, date_col, todays_date): 
    
    wb = load_workbook(filename)
    ws = wb.active    
    max_row = ws.max_row + 1
    
    
    ## Temp list to store row data empty strings as placeholders for formatting
    ##     columns in ODS file
    temp_list = ['','','','','','','','','','','','','','',]
    
    for r in range(starting_row,max_row):
        unit_id = ws.cell(row=r, column=unit_id_col).value
        vin = ws.cell(row=r, column=vin_col).value
        hub_reading = ws.cell(row=r, column=hub_col).value
        date = ws.cell(row=r, column=date_col).value
        
        if not hub_reading:
            hub_reading = ''
            
        if not date:
            date = ''
        
        
        temp_list[2] = unit_id
        temp_list[3] = unit_type
        temp_list[4] = vin
        
        try:
            temp_list[7] = hub_reading #####################################changed to string was int(hub_reading)
        except:
            temp_list[7] = 0
            
        if date:
            temp_list[5] = f'{date[6:10]}-{date[:2]}-{date[3:5]}'
        else:
            temp_list[5] = todays_date
        
        
        ## Carve out for mc_trailers.xlsx because TK reefers don't 
        ##     track to Mcleod, so adding them would be stupid
        ## Instead, we will be adding them from report generated
        ##    by TK Tracking. Super annoying...
        ## This block will hopefully be removed when we get rid of all the
        ##    TK units in the not too distant future
        
        if 'Reefer Hours Report - notification.xlsx' in filename:
            if unit_id.startswith('5') and int(unit_id) < 532111:
                temp_list = ['','','','','','','','','','','','','','',]
                continue
            else:
                if unit_id.startswith('5'):
                    unit_id = f'U{unit_id}'
                    temp_list[2] = unit_id
                    xlsx_rows.append(temp_list)
        else:
            xlsx_rows.append(temp_list)
        
        temp_list = ['','','','','','','','','','','','','','',]
        
    wb.close()
    

    

## Stupid func to get data from the TK Tracking sheet.
## Hope to be able to get rid of this eventually.  
## Carve out for '16 trailers that for some dumb reason still show up in
##     our equipment list.      
## Note, there is also 2 special cases for units 531775 and 531944 because their serial numbers
##      dont show on TK for some reason.


def collect_stupid_TK_data(filename, starting_row, unit_id_col, unit_type, vin_col, hub_col, date):        
    wb = load_workbook(filename)
    ws = wb.active    
    max_row = ws.max_row + 1
    
    
    ## Temp list to store row data
    temp_list = ['','','','','','','','','','','','','','',]
    
    for r in range(starting_row,max_row):
        unit_id = f'U{ws.cell(row=r, column=unit_id_col).value}'
        if unit_id == 'U531775':
            vin = '6001221139'
        elif unit_id == 'U531944':
            vin = '6001255267'
        else:
            vin = ws.cell(row=r, column=vin_col).value
        hub_reading = ws.cell(row=r, column=hub_col).value

        if unit_id.startswith("U5316"):
            continue
        else:
            temp_list[2] = unit_id
            temp_list[3] = unit_type
            temp_list[4] = vin
            temp_list[7] = hub_reading #################################changed to string was int(hub_reading)
            temp_list[5] = date
            
            xlsx_rows.append(temp_list)
            temp_list = ['','','','','','','','','','','','','','',]
        
    wb.close()
    

        
        
## Simple func to create ODS file from xlsx_list data
        
def create_combined_ODS(filename):
    data = OrderedDict()

    data.update({"Sheet1": xlsx_rows})
    
    save_data(filename, data)
        







        
        
        
# collect_xlsx_data(f'{ROOT_DIR}\\sheets\\mc_trucks.xlsx',3,2,3,4)
 
        
        
# collect_stupid_TK_data(f'{ROOT_DIR}\\sheets\\tk_hrs.xlsx', 5, 1, 11, today)        
   
# collect_xlsx_data(f'{ROOT_DIR}\\sheets\\mc_trailers.xlsx', 3, 2, 3, 4)   



# create_combined_ODS(f'{ROOT_DIR}\\temp\\Cetarisimport.ods')































